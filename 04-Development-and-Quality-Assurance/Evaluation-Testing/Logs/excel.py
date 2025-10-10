import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Constants
PACKET_SIZES = [128, 256, 384, 512, 640, 768, 896, 1024, 1152, 1280, 1408, 1536]
LEFT_SIZES = PACKET_SIZES[:6]
RIGHT_SIZES = PACKET_SIZES[6:]
BASE_DIR = "decoded_logs"

# Regex patterns
FLOW_BLOCK_REGEX = re.compile(r"(?s)Flow number:.*?Average loss-burst size.*?pkt")
METRICS = {
    "Flow #": r"Flow number:\s*(\d+)",
    "Average delay (s)": r"Average delay\s*=\s*([-\d.]+)",
    "Average jitter (s)": r"Average jitter\s*=\s*([-\d.]+)",
    "Bitrate (Kbit/s)": r"Average bitrate\s*=\s*([-\d.]+)",
    "Packet Loss": r"Packets dropped\s*=\s*\d+\s+\(([\d.]+)\s?%",
    "Minimum delay (s)": r"Minimum delay\s*=\s*([-\d.]+)",
    "Maximum delay (s)": r"Maximum delay\s*=\s*([-\d.]+)",
    "Average packet rate (pkt/s)": r"Average packet rate\s*=\s*([-\d.]+)"
}

# Extract flow metrics from a file
def extract_flows(filepath):
    with open(filepath, 'r') as f:
        content = f.read()
    flow_blocks = FLOW_BLOCK_REGEX.findall(content)
    flows = []
    for block in flow_blocks:
        data = {}
        for key, pattern in METRICS.items():
            match = re.search(pattern, block)
            if match:
                if key == "Flow #":
                    data[key] = int(match.group(1))
                else:
                    data[key] = float(match.group(1))
            else:
                data[key] = None
        flows.append(data)
    return pd.DataFrame(flows)

# Write flow data and average formulas to the worksheet
def write_table_to_excel(ws, start_col, start_row, size, df):
    ws.cell(row=start_row, column=start_col).value = f"{size}"
    headers = df.columns.tolist()
    for i, header in enumerate(headers):
        ws.cell(row=start_row + 1, column=start_col + i).value = header
    for row_idx, row in enumerate(df.values.tolist()):
        for col_idx, val in enumerate(row):
            ws.cell(row=start_row + 2 + row_idx, column=start_col + col_idx).value = val
    avg_row = start_row + 2 + len(df)
    ws.cell(row=avg_row, column=start_col).value = "Average"
    for col_idx in range(1, len(headers)):
        col_letter = get_column_letter(start_col + col_idx)
        data_start = start_row + 2
        data_end = data_start + len(df) - 1
        formula = f"=AVERAGE({col_letter}{data_start}:{col_letter}{data_end})"
        ws.cell(row=avg_row, column=start_col + col_idx).value = formula

# Prepare output Excel file
OUTPUT_FILE = "Data.xlsx"
wb = Workbook()
first_sheet = True

# IP/protocol combinations
combinations = [
    ("ipv4", "tcp", "IPv4 - TCP"),
    ("ipv4", "udp", "IPv4 - UDP"),
    ("ipv6", "tcp", "IPv6 - TCP"),
    ("ipv6", "udp", "IPv6 - UDP")
]

for ip_ver, proto, sheet_name in combinations:
    ws = wb.active if first_sheet else wb.create_sheet()
    ws.title = sheet_name
    first_sheet = False

    row_cursor_left = 2
    row_cursor_right = 2

    for size in LEFT_SIZES:
        filename = f"{ip_ver}_{proto}_{size}.txt"
        path = os.path.join(BASE_DIR, filename)
        if os.path.exists(path):
            df = extract_flows(path)
            write_table_to_excel(ws, start_col=1, start_row=row_cursor_left, size=size, df=df)
            row_cursor_left += len(df) + 5
        else:
            print(f"Missing file: {filename}")

    for size in RIGHT_SIZES:
        filename = f"{ip_ver}_{proto}_{size}.txt"
        path = os.path.join(BASE_DIR, filename)
        if os.path.exists(path):
            df = extract_flows(path)
            write_table_to_excel(ws, start_col=12, start_row=row_cursor_right, size=size, df=df)
            row_cursor_right += len(df) + 5
        else:
            print(f"Missing file: {filename}")

    # Place summary table headers (V-Z, row 3)
    summary_headers = ["Packet Size", "Average Delay", "Jitter", "Bitrate", "Packet Loss"]
    for i, header in enumerate(summary_headers):
        ws.cell(row=3, column=22 + i).value = header  # Columns V(22) to Z(26)

    # Reference cells for averages from your tables (adjusted row numbers)
    # These need to match where your average rows actually appear in the data tables.
    # Here I'm using your example row numbers — make sure these are correct for your dataset!
    delay_refs = ["B14", "B29", "B44", "B59", "B74", "B89", "M14", "M29", "M44", "M59", "M74", "M89"]
    jitter_refs = ["C14", "C29", "C44", "C59", "C74", "C89", "N14", "N29", "N44", "N59", "N74", "N89"]
    bitrate_refs = ["D14", "D29", "D44", "D59", "D74", "D89", "O14", "O29", "O44", "O59", "O74", "O89"]
    packetloss_refs = ["E14", "E29", "E44", "E59", "E74", "E89", "P14", "P29", "P44", "P59", "P74", "P89"]

    for i in range(12):
        row = 4 + i  # Start summary data from row 4
        ws.cell(row=row, column=22).value = PACKET_SIZES[i]  # Column V: Packet Size
        ws.cell(row=row, column=23).value = f"={delay_refs[i]}"     # Column W: Average Delay
        ws.cell(row=row, column=24).value = f"={jitter_refs[i]}"    # Column X: Jitter
        ws.cell(row=row, column=25).value = f"={bitrate_refs[i]}"   # Column Y: Bitrate
        ws.cell(row=row, column=26).value = f"={packetloss_refs[i]}"# Column Z: Packet Loss

# Save the workbook
wb.save(OUTPUT_FILE)
print(f"Done! Summary table added. Saved to: {OUTPUT_FILE}")
